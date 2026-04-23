import os
import random
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from readFiles.InterNetworkReader import InterNetworkReader
from readFiles.readVirtualNetwork import VirtualNetworkRequest
from readFiles.IntraNetworkReader import IntraNetworkReader
from readFiles.TransactionReader import TransactionReader

from algorithm.genetic_delay import GeneticDelaySolver, GreedyCPUDelaySolver, CentralityDelayGreedySolver


# -----------------------------------------------------------------------
# TrackingDelaySolver:
#   GeneticDelaySolver üzerine kurulu; her nesil boyunca kromozom ve
#   fitness geçmişini takip eder.
# -----------------------------------------------------------------------
class TrackingDelaySolver(GeneticDelaySolver):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._evaluated_chroms: list = []
        self._fitness_history:  list = []

    def calculate_fitness_v2(self, chromosome):
        self._evaluated_chroms.append(tuple(chromosome))
        return super().calculate_fitness_v2(chromosome)

    def _on_generation_end(self, _gen: int, best_fitness: float):
        self._fitness_history.append(best_fitness)

    def total_chrom_count(self):
        return len(self._evaluated_chroms)

    def unique_chrom_count(self):
        return len(set(self._evaluated_chroms))

    def get_gen_samples(self, population_size: int, n_samples: int = 3):
        result = []
        total  = len(self._evaluated_chroms)
        gen_no = 1
        start  = 0
        while start < total:
            end     = min(start + population_size, total)
            samples = list(self._evaluated_chroms[start : start + n_samples])
            result.append((gen_no, samples))
            start  = end
            gen_no += 1
        return result


# -----------------------------------------------------------------------
# Ayarlar
# -----------------------------------------------------------------------
networkType = "NSFNET"
folder      = f"topologies/{networkType}"
INTRA_NODE_COUNTS = [5]
#INTRA_NODE_COUNTS = [5, 6, 7, 8, 9, 10]

GA_POPULATION  = 60
GA_GENERATIONS = 5
GA_MUTATION    = 0.1
GA_SEED        = None

METHODS = [
    ("CPU (deterministik en-iyi)",  "cpu",       1.0),
    #("Rulet Tekerleği",              "roulette",  1.0),
    ("Rank Tabanlı",                 "rank",      1.0),
    #("Softmax / Boltzmann (T=1.0)", "softmax",   1.0),
    ("Q-Learning (e-greedy)",        "qlearning", 1.0),
]

GEN_SAMPLE_COUNT = 3


# -----------------------------------------------------------------------
# Veri yükleme
# -----------------------------------------------------------------------
def load_inter_network():
    directory = f"{folder}/internetwork/"
    for fname in os.listdir(directory):
        if fname.endswith('.txt'):
            return InterNetworkReader(os.path.join(directory, fname)), os.path.splitext(fname)[0]
    raise FileNotFoundError(f"Inter-network dosyası bulunamadı: {directory}")


def _parse_substrate_filename(fname):
    parts = fname.split("_")
    try:
        return {"snode": int(parts[1]), "slink": int(parts[2]), "scopy": int(parts[3])}
    except (IndexError, ValueError):
        return {"snode": "", "slink": "", "scopy": ""}


def load_intra_networks(num_domains, n_nodes):
    directory = f"{folder}/intranetwork/"
    list_file = f"{folder}/intra_domain_used_list.txt"

    with open(list_file) as f:
        all_lines = [
            line.strip() for line in f
            if line.strip() and not line.strip().startswith('#')
        ]

    prefix   = f"adjacency_{n_nodes}_"
    selected = [ln for ln in all_lines if ln.startswith(prefix)]

    if len(selected) < num_domains:
        raise FileNotFoundError(
            f"'{list_file}' icinde {n_nodes} node icin {num_domains} dosya bulunamadi "
            f"(bulunan: {len(selected)})."
        )

    selected = selected[:num_domains]

    print(f"[Liste Secim] Intra topoloji ({n_nodes} node, {num_domains} domain):")
    for i, f in enumerate(selected):
        print(f"  Domain {i:>2}: {f}")

    return IntraNetworkReader.load_intra_topology(directory, selected), selected


def load_transaction(n_nodes):
    directory = f"{folder}/transactions/"
    for fname in os.listdir(directory):
        if fname.endswith('.txt'):
            txt_name = os.path.splitext(fname)[0]
            prefix   = "_".join(txt_name.split('_')[:-1])
            return TransactionReader(
                os.path.join(directory, f"{prefix}_{n_nodes}.txt")
            )
    raise FileNotFoundError(f"Transaction dosyası bulunamadı: {directory}")


def load_virtual_requests():
    directory = f"{folder}/virtualrequests/"
    requests  = []
    filenames = []
    for fname in os.listdir(directory):
        if fname.endswith('.txt'):
            vr = VirtualNetworkRequest(os.path.join(directory, fname))
            if vr:
                requests.append(vr)
                filenames.append(os.path.splitext(fname)[0])
    if not requests:
        raise FileNotFoundError(f"Sanal ag istek dosyası bulunamadı: {directory}")
    return requests, filenames


# -----------------------------------------------------------------------
# Yol detayı yazdırma
# -----------------------------------------------------------------------
def yazYolDetaylari(yol_detaylari: list):
    print("\n  --- YOL DETAYLARI (DELAY) ---")
    if not yol_detaylari:
        print("  (yol bulunamadı)")
        return
    toplam_delay = 0
    for link in yol_detaylari:
        print(f"\n  {link['sanal_baglanti']}")
        if 'inter_domain_yolu' in link:
            print(f"    Inter-domain yolu: {link['inter_domain_yolu']}")
        for seg in link['segmentler']:
            tip = seg['tip']
            if tip == 'domain gecisi':
                er_delay = seg.get('hop', 0)
                print(f"    [Gecis] Domain {seg['kaynak_domain']} -> Domain {seg['hedef_domain']}  (delay={er_delay})")
                toplam_delay += er_delay
            elif tip == 'HATA':
                print(f"    [HATA]  {seg['mesaj']}")
            else:
                print(f"    [{tip}]  Domain {seg.get('domain','?')} | "
                      f"node {seg.get('baslangic_node','?')} -> node {seg.get('bitis_node','?')} | "
                      f"yol: {seg.get('yol','?')} | delay: {seg.get('hop','?')} | BW: {seg.get('bw_talebi','?')}")
                delay = seg.get('hop', 0)
                if isinstance(delay, (int, float)) and delay < 100000:
                    toplam_delay += delay
    print(f"\n  Toplam Delay: {toplam_delay}")


# Algoritma etiketi → Excel sütun öneki
KISA_AD = {
    "cpu":              "CPU",
    "roulette":         "Rulet",
    "rank":             "Rank",
    "softmax":          "Softmax",
    "qlearning":        "QL",
    "greedy_cpu":       "GreedyCPU",
    "greedy_closeness": "GreedyClose",
}

# Tek-geçişli greedy yöntemler — delay çözücüler
GREEDY_METHODS = [
    ("Greedy CPU (Delay)",       "greedy_cpu",       GreedyCPUDelaySolver,        None),
    ("Greedy Closeness (Delay)", "greedy_closeness", CentralityDelayGreedySolver, "closeness"),
]


def _ozet_hesapla(solver, best_chrom):
    """
    Bir algoritmanın en iyi kromozomu için:
      - toplam_delay : toplam delay değeri (delay versiyonunda hop yerine delay)
      - toplam_bw    : her link için delay × bw_talebi toplamı
      - bulunan_yol  : tüm segment yolları tek hücrede '-' ile birleştirilmiş
    döndürür.
    """
    if best_chrom is None:
        for gene_idx in range(solver.num_genes):
            cpu_req = solver.cpu_demand_VirtualNetwork[gene_idx]
            d1, d2  = solver.candidateDomains[gene_idx]
            herhangi_yeterli = any(
                cap >= cpu_req
                for d in (d1, d2)
                for cap in solver.cpu_value_all_intra_networks[d]
            )
            if not herhangi_yeterli:
                return "Yetersiz CPU", "Yetersiz CPU", "Yetersiz CPU"
        return "cozum yok", "cozum yok", "cozum yok"

    for gene_idx, gene in enumerate(best_chrom):
        d, node = solver._parse_gene(gene)
        cpu_cap = solver.cpu_value_all_intra_networks[d][node]
        cpu_req = solver.cpu_demand_VirtualNetwork[gene_idx]
        if cpu_cap < cpu_req:
            return "Yetersiz CPU", "Yetersiz CPU", "Yetersiz CPU"

    yol_detaylari = solver.trace_solution(best_chrom)
    toplam_delay = 0
    toplam_bw    = 0
    yol_parcalari = []
    bw_hatasi = False

    for link in yol_detaylari:
        bw_talebi = link.get("bw_talebi", 0) or 0
        for seg in link.get("segmentler", []):
            tip  = seg.get("tip", "")
            delay = seg.get("hop", 0)
            if tip == "HATA":
                bw_hatasi = True
                yol_parcalari.append("HATA")
                continue
            if isinstance(delay, (int, float)) and delay >= 100000:
                bw_hatasi = True
                continue
            if tip == "domain gecisi":
                yol_parcalari.append(
                    f"D{seg.get('kaynak_domain','')}->D{seg.get('hedef_domain','')}"
                )
                toplam_delay += delay
                toplam_bw    += delay * bw_talebi
                continue
            yol = seg.get("yol") or []
            if isinstance(yol, list):
                yol_parcalari.append("-".join(str(n) for n in yol))
            else:
                yol_parcalari.append(str(yol))
            toplam_delay += delay
            toplam_bw    += delay * bw_talebi

    if bw_hatasi:
        return "Yetersiz BW", "Yetersiz BW", "Yetersiz BW"

    bulunan_yol = " | ".join(yol_parcalari)
    return toplam_delay, toplam_bw, bulunan_yol


def _parse_vr_filename(fname):
    """
    'virtual_5_5_15_1' → {'vn': 5, 'bw': 5, 'cpu': 15, 'copy': 1}
    Format: virtual_{VNsayisi}_{BW}_{CPU}_{copy}
    """
    parts = fname.split("_")
    try:
        return {
            "dosya": fname,
            "vn":    int(parts[1]),
            "bw":    int(parts[2]),
            "cpu":   int(parts[3]),
            "copy":  int(parts[4]),
        }
    except (IndexError, ValueError):
        return {"dosya": fname, "vn": "", "bw": "", "cpu": "", "copy": ""}


# -----------------------------------------------------------------------
# Excel kayıt — delay versiyonu
# -----------------------------------------------------------------------
def olustur_excel_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sonuclar"

    baslik_fill = PatternFill("solid", fgColor="BFBFBF")
    baslik_font = Font(bold=True)

    algo_basliklar = []
    for _, key, _, _ in GREEDY_METHODS:
        on = KISA_AD.get(key, key)
        algo_basliklar += [
            f"{on}_Delay",
            f"{on}_BW",
            f"{on}_Path",
            f"{on}_Chromosome",
        ]
    for _, mode, _ in METHODS:
        on = KISA_AD.get(mode, mode)
        algo_basliklar += [
            f"GA_{on}_Delay",
            f"GA_{on}_BW",
            f"GA_{on}_Path",
            f"GA_{on}_Chromosome",
        ]

    meta = ["VNode", "VBW", "VCPU", "VCopy", "SNode", "SLink", "SCopy", "NodePerISP"]

    ws.append(algo_basliklar + meta)
    for cell in ws[1]:
        cell.font      = baslik_font
        cell.fill      = baslik_fill
        cell.alignment = Alignment(horizontal="center")

    ws2 = wb.create_sheet(title="Fİltreli")
    ozet_basliklar = [b for b in algo_basliklar if not b.endswith("_Path") and not b.endswith("_Chromosome")]
    ws2.append(ozet_basliklar + meta)
    for cell in ws2[1]:
        cell.font      = baslik_font
        cell.fill      = baslik_fill
        cell.alignment = Alignment(horizontal="center")

    return wb


def kaydet_excel(wb, greedy_sonuclar: list, greedy_solvers: list,
                 sonuclar: list, solvers: list, vr_fname: str,
                 substrate_fname: str = "", node_per_isp: int = 0):
    ws   = wb.active
    dosya = "sonuclar_delay.xlsx"

    satir = []
    for (_, best_chrom, _), solver in zip(greedy_sonuclar, greedy_solvers):
        toplam_delay, toplam_bw, bulunan_yol = _ozet_hesapla(solver, best_chrom)
        kromozom_str = str(best_chrom) if best_chrom is not None else "cozum yok"
        satir += [toplam_delay, toplam_bw, bulunan_yol, kromozom_str]
    for (_, best_chrom, _), solver in zip(sonuclar, solvers):
        toplam_delay, toplam_bw, bulunan_yol = _ozet_hesapla(solver, best_chrom)
        kromozom_str = str(best_chrom) if best_chrom is not None else "cozum yok"
        satir += [toplam_delay, toplam_bw, bulunan_yol, kromozom_str]

    vr_info  = _parse_vr_filename(vr_fname)
    sub_info = _parse_substrate_filename(substrate_fname)
    meta_satir = [vr_info["vn"], vr_info["bw"], vr_info["cpu"], vr_info["copy"],
                  sub_info["snode"], sub_info["slink"], sub_info["scopy"],
                  node_per_isp]

    ws.append(satir + meta_satir)

    ws2 = wb["Fİltreli"]
    ozet_satir = []
    for i in range(0, len(satir), 4):
        ozet_satir += satir[i:i+2]
    ws2.append(ozet_satir + meta_satir)

    for ws_iter in (ws, ws2):
        for col in ws_iter.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=8)
            ws_iter.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(dosya)
    print(f"\n  [Excel guncellendi: {dosya}  →  satir eklendi: {vr_fname}]")


# -----------------------------------------------------------------------
# Fitness eğrisi grafiği
# -----------------------------------------------------------------------
def ciz_fitness_egrisi(solvers: list, labels: list, vr_idx: int):
    plt.figure(figsize=(10, 6))

    renkler  = ['tab:blue', 'tab:orange', 'tab:green', 'tab:red', 'tab:purple']
    isaretler = ['o', 's', '^', 'D', 'x']

    for solver, label, renk, isaret in zip(solvers, labels, renkler, isaretler):
        if not solver._fitness_history:
            continue
        nesiller = list(range(1, len(solver._fitness_history) + 1))
        plt.plot(nesiller, solver._fitness_history,
                 label=label, color=renk, marker=isaret,
                 linewidth=2, markersize=6)

    plt.title(f"Nesil Boyunca En Iyi Fitness (Delay)  —  Sanal Ag Istegi #{vr_idx + 1}")
    plt.xlabel("Nesil")
    plt.ylabel("En Iyi Fitness (Toplam Delay)")
    plt.legend(loc="upper right")
    plt.grid(True, linestyle='--', alpha=0.5)
    plt.tight_layout()

    dosya_adi = f"fitness_egrisi_delay_vr{vr_idx + 1}.png"
    plt.savefig(dosya_adi)
    print(f"\n  [Grafik kaydedildi: {dosya_adi}]")
    plt.show()


# -----------------------------------------------------------------------
# Nesil boyunca kromozom çeşitliliği yazdırma
# -----------------------------------------------------------------------
def yazNesil_Kromozomlar(solver: TrackingDelaySolver, label: str, population_size: int):
    print(f"\n  [{label}]")
    gen_samples = solver.get_gen_samples(population_size, GEN_SAMPLE_COUNT)
    toplam = len(gen_samples)
    if toplam == 0:
        return

    orta = toplam // 2

    gosterilecekler = []
    gosterilecekler.append(gen_samples[0])
    if toplam > 2:
        gosterilecekler.append(gen_samples[orta])
    gosterilecekler.append(gen_samples[-1])

    isimler = {gen_samples[0][0]: "Bas", gen_samples[-1][0]: "Son"}
    if toplam > 2:
        isimler[gen_samples[orta][0]] = "Orta"

    for gen_no, samples in gosterilecekler:
        etiket = isimler.get(gen_no, "")
        print(f"    -- Nesil {gen_no} ({etiket}) --")
        for idx, chrom in enumerate(samples):
            print(f"      #{idx+1}: {list(chrom)}")


# -----------------------------------------------------------------------
# Tek-geçişli greedy yöntem çalıştırma
# -----------------------------------------------------------------------
def run_greedy_method(label, solver_class, centrality_method,
                      allTransaction, edgeRouter, interNetwork,
                      intraTopologies, virtualRequests):

    print(f"\n{'-' * 70}")
    print(f"  YONTEM: {label}")
    print(f"{'-' * 70}")

    if centrality_method is not None:
        solver = solver_class(
            allTransaction, edgeRouter, interNetwork, intraTopologies, virtualRequests,
            method=centrality_method,
        )
    else:
        solver = solver_class(
            allTransaction, edgeRouter, interNetwork, intraTopologies, virtualRequests,
        )

    chrom, fitness, path_details = solver.solve()

    print(f"\n  En iyi fitness (delay) : {fitness}")
    print(f"  En iyi kromozom        : {chrom}")
    if chrom is not None:
        yazYolDetaylari(path_details)

    return chrom, fitness, solver


# -----------------------------------------------------------------------
# Tek bir GA yöntemi çalıştırma
# -----------------------------------------------------------------------
def run_method(label, selection_mode, softmax_temp,
               allTransaction, edgeRouter, interNetwork,
               intraTopologies, virtualRequests):

    print(f"\n{'-' * 70}")
    print(f"  YONTEM: {label}")
    print(f"{'-' * 70}")

    solver = TrackingDelaySolver(
        allTransaction, edgeRouter, interNetwork, intraTopologies, virtualRequests,
        selection_mode=selection_mode,
        softmax_temperature=softmax_temp,
    )
    result = solver.run(
        population_size=GA_POPULATION,
        generations=GA_GENERATIONS,
        mutation_rate=GA_MUTATION,
        seed=GA_SEED,
    )
    if result is None:
        best_chrom, best_fitness = None, float('inf')
    else:
        best_chrom, best_fitness = result

    total  = solver.total_chrom_count()
    unique = solver.unique_chrom_count()

    print(f"\n  En iyi fitness (delay) : {best_fitness}")
    print(f"  En iyi kromozom        : {best_chrom}")
    print(f"  Toplam degerlendirilen: {total}  |  Benzersiz: {unique}  |  "
          f"Tekrar: {total - unique} ({(total - unique) / max(total, 1) * 100:.1f}%)")

    if selection_mode == 'qlearning' and solver.q_table:
        print(f"\n  Q-Tablosu boyutu : {len(solver.q_table)} girdi")
        print(f"  Son epsilon      : {solver.ql_epsilon:.4f}  "
              f"(baslangic=0.8 → her nesil x{solver.ql_epsilon_decay} → min={solver.ql_epsilon_min})")
        en_iyi = sorted(solver.q_table.items(), key=lambda x: x[1], reverse=True)[:5]
        print("  En yuksek 5 Q-degeri  (gen, domain, doluluk, node) -> Q:")
        for (gi, di, dol, ni), qv in en_iyi:
            print(f"    gen={gi}  domain={di}  doluluk={dol}  node={ni}  ->  Q={qv:.4f}")

    if best_chrom is not None:
        yazYolDetaylari(solver.trace_solution(best_chrom))

    return best_chrom, best_fitness, solver


# -----------------------------------------------------------------------
# Başlangıç sabitleri — modül yüklenirken bir kez okunur
# Domain ID'leri tüm dosyalarda 0-tabanlıdır:
#   VR istekleri : 0 = 1. domain, NUM_DOMAINS-1 = son domain
#   Transaction  : Ingress/Egress "000xxx" → domain 0, "013xxx" → domain 13
#   Edge router  : aynı 0-tabanlı kural
# NUM_DOMAINS inter-network komşuluk matrisinin satır sayısından belirlenir.
# -----------------------------------------------------------------------
interNetwork, SUBSTRATE_FNAME = load_inter_network()
NUM_DOMAINS = interNetwork.get_numberOfInterNodes()


# -----------------------------------------------------------------------
# Ana fonksiyon
# -----------------------------------------------------------------------
def main():
    print(f"[0] Inter-network yuklendi: {SUBSTRATE_FNAME}  |  NUM_DOMAINS={NUM_DOMAINS}")

    print("[0] Sanal ag istekleri yukleniyor...")
    virtual_requests_list, vr_filenames = load_virtual_requests()

    wb = olustur_excel_workbook()

    for n_nodes in INTRA_NODE_COUNTS:
        print(f"\n{'*' * 70}")
        print(f"  INTRA NODE SAYISI: {n_nodes}  [DELAY MODU]")
        print(f"{'*' * 70}")

        print(f"[1] Intra-network yukleniyor ({NUM_DOMAINS} domain, n={n_nodes})...")
        intraTopologies, intraNameList = load_intra_networks(NUM_DOMAINS, n_nodes)

        print(f"[2] Transaction yukleniyor (n={n_nodes})...")
        allTransaction, edgeRouter = load_transaction(n_nodes)

        for vr_idx, (virtualRequests, vr_fname) in enumerate(zip(virtual_requests_list, vr_filenames)):
            print(f"\n{'#' * 70}")
            print(f"  Sanal Ag Istegi #{vr_idx + 1}  |  nodePerISP={n_nodes}  [DELAY MODU]")
            print(f"{'#' * 70}")

            greedy_sonuclar = []
            greedy_solvers  = []
            sonuclar = []
            solvers  = []

            # ---- Greedy yöntemler (delay) ----
            for label, _, solver_class, cent_method in GREEDY_METHODS:
                chrom, fitness, solver = run_greedy_method(
                    label, solver_class, cent_method,
                    allTransaction, edgeRouter, interNetwork,
                    intraTopologies, virtualRequests,
                )
                greedy_sonuclar.append((label, chrom, fitness))
                greedy_solvers.append(solver)

            # ---- GA yöntemleri (delay) ----
            for label, mode, temp in METHODS:
                chrom, fitness, solver = run_method(
                    label, mode, temp,
                    allTransaction, edgeRouter, interNetwork,
                    intraTopologies, virtualRequests,
                )
                sonuclar.append((label, chrom, fitness))
                solvers.append(solver)

            # ---------------------------------------------------------------
            # KARSILASTIRMA TABLOSU
            # ---------------------------------------------------------------
            print(f"\n\n{'=' * 70}")
            print(f"  KARSILASTIRMA TABLOSU (DELAY)  —  VR #{vr_idx + 1}  |  nodePerISP={n_nodes}")
            print(f"{'=' * 70}")
            print(f"  {'Yontem':<40} {'Fitness':>10}  {'Toplam':>7}  {'Benzersiz':>9}  {'Oran':>6}")
            print(f"  {'-' * 40}  {'-' * 10}  {'-' * 7}  {'-' * 9}  {'-' * 6}")

            tum_sonuclar = greedy_sonuclar + sonuclar
            best_fit = min((f for _, _, f in tum_sonuclar if f is not None), default=float('inf'))

            for label, chrom, fitness in greedy_sonuclar:
                marker    = " << EN IYI" if fitness == best_fit else ""
                fit_str   = str(fitness) if fitness is not None else "cozum yok"
                chrom_str = str(chrom)   if chrom   is not None else "cozum yok"
                print(f"  {label:<40} {fit_str:>10}  {'1':>7}  {'1':>9}  {'100%':>6}{marker}  {chrom_str}")

            for (label, chrom, fitness), solver in zip(sonuclar, solvers):
                t    = solver.total_chrom_count()
                u    = solver.unique_chrom_count()
                oran = f"{u / max(t, 1) * 100:.1f}%"
                marker    = " << EN IYI" if fitness == best_fit else ""
                fit_str   = str(fitness) if fitness is not None else "cozum yok"
                chrom_str = str(chrom)   if chrom   is not None else "cozum yok"
                print(f"  {label:<40} {fit_str:>10}  {t:>7}  {u:>9}  {oran:>6}{marker}  {chrom_str}")

            # ---------------------------------------------------------------
            # EXCEL KAYIT
            # ---------------------------------------------------------------
            kaydet_excel(wb, greedy_sonuclar, greedy_solvers, sonuclar, solvers, vr_fname,
                         substrate_fname=SUBSTRATE_FNAME,
                         node_per_isp=n_nodes)

            # ---------------------------------------------------------------
            # TUM NESILLER BOYUNCA KROMOZU CESITLILIGI
            # ---------------------------------------------------------------
            print(f"\n{'=' * 70}")
            print("  TUM NESILLER BOYUNCA KROMOZU CESITLILIGI")
            print(f"{'=' * 70}")

            for (label, _, _), solver in zip(METHODS, solvers):
                yazNesil_Kromozomlar(solver, label, GA_POPULATION)

        print(f"\nKullanilan Intra Topolojiler (n={n_nodes}, {len(intraNameList)} domain):")
        for idx, name in enumerate(intraNameList):
            print(f"  Domain {idx:>2}: {name}")


if __name__ == "__main__":
    main()
