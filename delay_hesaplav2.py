"""
delay_hesapla.py

Excel sonuç dosyasındaki kromozom bilgisini kullanarak her VNR için
uç-uç gecikme (end-to-end delay) hesaplar; her algoritmanın _BW sütununun
hemen arkasına _Delay sütunu olarak ekler ve orijinali bozmadan
_delay.xlsx uzantılı kopyayı kaydeder.

Gecikme hesaplama modeli
─────────────────────────
Sanal link (vi, vj):
  • Aynı domain   → intra delay_matrix üzerinde Dijkstra
  • Farklı domain → Σ( intra_delay(cur→er_giris) + er.maxDelay ) her domain
                    geçişi için; son domaininde intra_delay(er_cikis→hedef)

VNR gecikmesi = BFS ile sanal topolojide her kaynak düğümden erişilebilen
                tüm hedef düğümlere birikimli fiziksel gecikmelerin maksimumu.

Örnekler:
  v0-v1, v0-v2, v0-v3, v0-v4  → max(d01, d02, d03, d04)
  v0-v1, v0-v2, v2-v3, v2-v4  → max(d01, d02+d23, d02+d24)

Kullanım:
    python delay_hesapla.py                          # varsayılan INPUT_EXCEL
    python delay_hesapla.py sonuclar_US_isp6.xlsx
"""

import os
import sys
import ast
import math

import networkx as nx
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from readFiles.IntraNetworkReader import IntraNetworkReader
from readFiles.InterNetworkReader import InterNetworkReader
from readFiles.TransactionReader import TransactionReader
from readFiles.readVirtualNetwork import VirtualNetworkRequest

# ─── AYARLAR ─────────────────────────────────────────────────────────────────
INPUT_EXCEL    = "sonuclar_US_guncel_isp10_3000req.xlsx"
NETWORK_FOLDER = "topologies/guncel/USNET"   # NSFNET veya USNET
# ─────────────────────────────────────────────────────────────────────────────

ALGOS = [
    "GreedyCPU", "GreedyCPUBW", "GreedyBWSort", "GreedyDegSort",
    "GreedyClose", "GreedyCloseBW", "GreedyCloseDeg",
    "GA_CPU", "GA_Rank", "GA_QL",
]

INF = math.inf


# ─── Topoloji yardımcıları ────────────────────────────────────────────────────

def load_topology(network_folder, n_nodes):
    """Intra/inter/transaction dosyalarını yükler; delay graflarını döndürür."""
    # Inter network
    inter_dir = os.path.join(network_folder, "internetwork")
    inter_fname = next(f for f in os.listdir(inter_dir) if f.endswith(".txt"))
    inter_net = InterNetworkReader(os.path.join(inter_dir, inter_fname))
    num_domains = inter_net.get_numberOfInterNodes()

    # Intra networks
    list_file = os.path.join(network_folder, "intra_domain_used_list.txt")
    with open(list_file) as fh:
        all_lines = [
            ln.strip() for ln in fh
            if ln.strip() and not ln.strip().startswith("#")
        ]
    prefix   = f"adjacency_{n_nodes}_"
    selected = [ln for ln in all_lines if ln.startswith(prefix)][:num_domains]
    intra_dir   = os.path.join(network_folder, "intranetwork")
    intra_topos = IntraNetworkReader.load_intra_topology(intra_dir, selected)

    # Transactions + edge routers
    trans_dir   = os.path.join(network_folder, "transactions")
    trans_fname = next(
        f for f in os.listdir(trans_dir) if f.endswith(f"_{n_nodes}.txt")
    )
    all_trans, edge_routers = TransactionReader(
        os.path.join(trans_dir, trans_fname)
    )

    return intra_topos, all_trans, edge_routers, num_domains


def build_intra_delay_graphs(intra_topos):
    """Her domain için gecikme ağırlıklı NetworkX grafı döndürür."""
    graphs = []
    for topo in intra_topos:
        G   = nx.Graph()
        adj = topo.adjacency_matrix
        dly = topo.delay_matrix
        n   = len(adj)
        for i in range(n):
            for j in range(i + 1, n):
                if adj[i][j] == 1:
                    G.add_edge(i, j, delay=dly[i][j])
        graphs.append(G)
    return graphs


def build_inter_domain_graph(edge_routers, num_domains):
    """Domain bağlantı grafı (hop sayısı = 1 her kenar için)."""
    G = nx.Graph()
    G.add_nodes_from(range(num_domains))
    for er in edge_routers:
        G.add_edge(er.edgeDomainIngress, er.edgeDomainEgress)
    return G


# ─── Gecikme hesaplama ────────────────────────────────────────────────────────

def _dijkstra_delay(G, src, dst):
    """Aynı domain içinde iki düğüm arasındaki en kısa gecikme.
    Yol yoksa INF döner (önceden sessizce 0 dönüyordu)."""
    try:
        return nx.dijkstra_path_length(G, src, dst, weight="delay")
    except (nx.NetworkXNoPath, nx.NodeNotFound):
        return INF


def _find_edge_router(cur_domain, next_domain, edge_routers):
    """cur_domain ↔ next_domain bağlayan ilk edge-router'ı döndürür."""
    for er in edge_routers:
        if er.edgeDomainIngress == cur_domain and er.edgeDomainEgress == next_domain:
            return er, er.edgeNodeIngress, er.edgeDomainEgress, er.edgeNodeEgress
        if er.edgeDomainEgress == cur_domain and er.edgeDomainIngress == next_domain:
            return er, er.edgeNodeEgress, er.edgeDomainIngress, er.edgeNodeIngress
    return None, None, None, None


def _bc_delay(cur_domain, cur_node, target_node, all_trans):
    """BC transaction varsa min-delay'ini döndürür; yoksa None."""
    best = None
    for t in all_trans:
        if t.edgeDomainEgress != cur_domain:
            continue
        if (
            (t.edgeNodeIngress == cur_node  and t.edgeNodeEgress == target_node) or
            (t.edgeNodeIngress == target_node and t.edgeNodeEgress == cur_node)
        ):
            if best is None or t.maxDelay < best:
                best = t.maxDelay
    return best


def compute_link_delay(
    d_i, n_i, d_j, n_j,
    intra_graphs, edge_routers, inter_G, all_trans, er_node_set,
):
    """
    Sanal link (d_i,n_i) → (d_j,n_j) için fiziksel gecikmeyi hesaplar.
    Aynı domain: Dijkstra (delay_matrix ağırlıklı).
    Farklı domain: inter-domain hop yolu üzerinden gecikme toplamı.
    Hata durumlarında INF döner (önceden sessizce 0 dönüyordu).
    """
    if d_i == d_j:
        return _dijkstra_delay(intra_graphs[d_i], n_i, n_j)

    try:
        inter_path = nx.shortest_path(inter_G, source=d_i, target=d_j)
    except (nx.NetworkXNoPath, nx.NodeNotFound):
        return INF

    total  = 0
    cur_d  = d_i
    cur_n  = n_i

    for step in range(len(inter_path) - 1):
        er, er_ing, next_d, next_n = _find_edge_router(
            inter_path[step], inter_path[step + 1], edge_routers
        )
        if er is None:
            return INF

        # Intra-domain: cur_n → er giriş noktası
        if (cur_d, cur_n) in er_node_set:
            # BC transaction dene
            bc = _bc_delay(cur_d, cur_n, er_ing, all_trans)
            total += bc if bc is not None else _dijkstra_delay(intra_graphs[cur_d], cur_n, er_ing)
        else:
            total += _dijkstra_delay(intra_graphs[cur_d], cur_n, er_ing)

        # Erken çıkış: alt yol zaten erişilemezse devam etmeye gerek yok
        if total == INF:
            return INF

        # Domain sınırı geçiş gecikmesi
        total += er.maxDelay

        cur_d = next_d
        cur_n = next_n

    # Son domain: cur_n → hedef düğüm
    if (cur_d, cur_n) in er_node_set:
        bc = _bc_delay(cur_d, cur_n, n_j, all_trans)
        total += bc if bc is not None else _dijkstra_delay(intra_graphs[cur_d], cur_n, n_j)
    else:
        total += _dijkstra_delay(intra_graphs[cur_d], cur_n, n_j)

    return total


def compute_vnr_delay(
    chrom_str, adj_matrix,
    intra_graphs, edge_routers, inter_G, all_trans, er_node_set,
):
    """
    Kromozon ve sanal topoloji kullanarak VNR gecikmesini hesaplar.

    BFS ile sanal topoloji üzerinde her düğümden birikimli gecikmelerin
    maksimumu döner (gecikme çapı / delay diameter).

    Dönüş değerleri:
      • float           → başarılı hesaplama (gecikme değeri)
      • "Hata: parse"   → kromozon parse edilemedi
      • "Hata: yol yok" → en az bir sanal link için fiziksel yol bulunamadı
    """
    try:
        chrom = ast.literal_eval(chrom_str)
    except Exception:
        return "Hata: parse"

    if not isinstance(chrom, list):
        return "Hata: parse"

    n_vn = len(adj_matrix)

    # Kromozon → (domain, node) eşlemesi
    assignments = {}
    for i, gene in enumerate(chrom):
        try:
            d, node = map(int, str(gene).split("-"))
            assignments[i] = (d, node)
        except Exception:
            return "Hata: parse"

    # Her sanal link için fiziksel gecikme hesapla
    G_V = nx.Graph()
    G_V.add_nodes_from(range(n_vn))

    for vi in range(n_vn):
        for vj in range(vi + 1, n_vn):
            if adj_matrix[vi][vj] <= 0:
                continue
            d_i, n_i = assignments[vi]
            d_j, n_j = assignments[vj]
            phys_d = compute_link_delay(
                d_i, n_i, d_j, n_j,
                intra_graphs, edge_routers, inter_G, all_trans, er_node_set,
            )
            if phys_d == INF:
                return "Hata: yol yok"
            G_V.add_edge(vi, vj, delay=phys_d)

    if G_V.number_of_edges() == 0:
        return 0

    # BFS: v0 (sanal düğüm 0) kaynaktan tüm düğümlere birikimli gecikmelerin maksimumu.
    # Sanal ağ bir tree yapısı; v0 kök kabul edilir.
    max_delay = 0
    bfs_paths = nx.single_source_shortest_path(G_V, 0)
    for target, path in bfs_paths.items():
        if target == 0 or len(path) < 2:
            continue
        path_delay = sum(
            G_V[path[k]][path[k + 1]]["delay"]
            for k in range(len(path) - 1)
        )
        if path_delay > max_delay:
            max_delay = path_delay

    return max_delay


# ─── Ana fonksiyon ────────────────────────────────────────────────────────────

def main(input_excel=None):
    src = input_excel or INPUT_EXCEL
    dst = src.replace(".xlsx", "_delay.xlsx")
    if dst == src:
        dst = src[:-5] + "_delay.xlsx"

    print(f"Girdi : {src}")
    print(f"Çıktı : {dst}")

    # Excel oku
    wb = openpyxl.load_workbook(src)
    ws = wb["Sonuclar"]

    headers = [cell.value for cell in ws[1]]

    # NodePerISP
    nisp_col = headers.index("NodePerISP") + 1
    n_nodes  = ws.cell(2, nisp_col).value
    print(f"NodePerISP = {n_nodes}")

    # Topoloji yükle
    print("Topoloji yükleniyor...")
    intra_topos, all_trans, edge_routers, num_domains = load_topology(
        NETWORK_FOLDER, n_nodes
    )
    intra_graphs = build_intra_delay_graphs(intra_topos)
    inter_G      = build_inter_domain_graph(edge_routers, num_domains)

    # Edge-router düğüm seti (BC transaction tespiti için)
    er_node_set = set()
    for er in edge_routers:
        er_node_set.add((er.edgeDomainIngress, er.edgeNodeIngress))
        er_node_set.add((er.edgeDomainEgress,  er.edgeNodeEgress))

    # VR önbelleği
    vr_dir   = os.path.join(NETWORK_FOLDER, "virtualrequests")
    vr_cache = {}

    def get_adj(vnode, vbw, vcpu, vcopy):
        key = (vnode, vbw, vcpu, vcopy)
        if key not in vr_cache:
            try:
                fpath = os.path.join(vr_dir, f"virtual_{vnode}_{vbw}_{vcpu}_{vcopy}.txt")
                vr_cache[key] = VirtualNetworkRequest(fpath).adjacency_matrix
            except Exception:
                vr_cache[key] = None
        return vr_cache[key]

    # Sütun indeksleri
    vnode_col = headers.index("VNode")  + 1
    vbw_col   = headers.index("VBW")   + 1
    vcpu_col  = headers.index("VCPU")  + 1
    vcopy_col = headers.index("VCopy") + 1

    chrom_cols = {}
    bw_cols    = {}
    for algo in ALGOS:
        bh = f"{algo}_BW"
        ch = f"{algo}_Chromosome"
        if bh in headers:
            bw_cols[algo]    = headers.index(bh) + 1
        if ch in headers:
            chrom_cols[algo] = headers.index(ch) + 1

    total_rows = ws.max_row - 1
    print(f"{total_rows} satır işlenecek...")

    # ── 1. Tüm gecikmeler önce hesaplanır (sütun indeksleri değişmeden önce) ──
    delays = {}   # {(row_idx, algo): değer}
    error_counts = {"parse": 0, "yol_yok": 0, "yetersiz_bw": 0, "hata": 0}

    for row_idx in range(2, ws.max_row + 1):
        if (row_idx - 1) % 200 == 0:
            print(f"  [{row_idx - 1}/{total_rows}] hesaplanıyor...")

        vnode = ws.cell(row_idx, vnode_col).value
        vbw   = ws.cell(row_idx, vbw_col).value
        vcpu  = ws.cell(row_idx, vcpu_col).value
        vcopy = ws.cell(row_idx, vcopy_col).value
        adj   = get_adj(vnode, vbw, vcpu, vcopy)

        for algo in ALGOS:
            chrom_val = ws.cell(row_idx, chrom_cols[algo]).value if algo in chrom_cols else None

            if chrom_val is None or str(chrom_val).strip() == "Yetersiz BW" or adj is None:
                delays[(row_idx, algo)] = "Yetersiz BW"
                error_counts["yetersiz_bw"] += 1
                continue

            result = compute_vnr_delay(
                chrom_val, adj,
                intra_graphs, edge_routers, inter_G, all_trans, er_node_set,
            )

            if result is None:
                delays[(row_idx, algo)] = "Hata"
                error_counts["hata"] += 1
            elif isinstance(result, str):
                # "Hata: parse" veya "Hata: yol yok"
                delays[(row_idx, algo)] = result
                if "parse" in result:
                    error_counts["parse"] += 1
                elif "yol yok" in result:
                    error_counts["yol_yok"] += 1
            else:
                delays[(row_idx, algo)] = result

    # Hata özeti
    print("\n─── Hata Özeti ───")
    print(f"  Yetersiz BW   : {error_counts['yetersiz_bw']}")
    print(f"  Parse hatası  : {error_counts['parse']}")
    print(f"  Yol bulunamadı: {error_counts['yol_yok']}")
    print(f"  Diğer hata    : {error_counts['hata']}")
    print("──────────────────\n")

    print("Gecikme sutunlari ekleniyor (Sonuclar)...")

    # ── 2a. Sonuclar: sağdan sola ekle ───────────────────────────────────────
    for algo in reversed(ALGOS):
        if algo not in bw_cols:
            continue
        insert_pos = bw_cols[algo] + 1

        ws.insert_cols(insert_pos)
        ws.cell(1, insert_pos).value = f"{algo}_Delay"

        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row_idx, insert_pos).value = delays.get((row_idx, algo), "Hata")

    # ── 2b. Filtreli: aynı delays sözlüğünü kullan, sağdan sola ekle ─────────
    print("Gecikme sutunlari ekleniyor (Filtreli)...")

    ws_f        = wb[wb.sheetnames[1]]
    headers_f   = [cell.value for cell in ws_f[1]]
    bw_cols_f   = {}
    for algo in ALGOS:
        bh = f"{algo}_BW"
        if bh in headers_f:
            bw_cols_f[algo] = headers_f.index(bh) + 1

    for algo in reversed(ALGOS):
        if algo not in bw_cols_f:
            continue
        insert_pos = bw_cols_f[algo] + 1

        ws_f.insert_cols(insert_pos)
        ws_f.cell(1, insert_pos).value = f"{algo}_Delay"

        for row_idx in range(2, ws_f.max_row + 1):
            ws_f.cell(row_idx, insert_pos).value = delays.get((row_idx, algo), "Hata")

    wb.save(dst)
    print(f"Kaydedildi: {dst}")


if __name__ == "__main__":
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    main(arg)